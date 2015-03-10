# coding: utf-8
from __future__ import unicode_literals
from .app.models import Person, Region
from attest import assert_hook, raises, Tests  # pylint: disable=W0611
from contextlib import contextmanager
from django_attest import queries, settings, TestContext, translation
import django_tables2 as tables
from django_tables2.config import RequestConfig
from django_tables2.utils import build_request
import django
from django.core.exceptions import ImproperlyConfigured
from django.template import Template, RequestContext, Context
from django.utils.translation import ugettext_lazy
from django.utils.safestring import mark_safe
try:
    from urlparse import parse_qs
except ImportError:
    from urllib.parse import parse_qs
import lxml.etree
import lxml.html
try:
    from cStringIO import StringIO
except ImportError:
    from StringIO import StringIO
import six


def parse(html):
    return lxml.etree.fromstring(html)


def attrs(xml):
    """
    Helper function that returns a dict of XML attributes, given an element.
    """
    return lxml.html.fromstring(xml).attrib


database = contextmanager(TestContext())
templates = Tests()


class CountryTable(tables.Table):
    name = tables.Column()
    capital = tables.Column(orderable=False,
                            verbose_name=ugettext_lazy("capital"))
    population = tables.Column(verbose_name='population size')
    currency = tables.Column(visible=False)
    tld = tables.Column(visible=False, verbose_name='domain')
    calling_code = tables.Column(accessor='cc',
                                 verbose_name='phone ext.')


MEMORY_DATA = [
    {'name': 'Germany', 'capital': 'Berlin', 'population': 83,
     'currency': 'Euro (€)', 'tld': 'de', 'cc': 49},
    {'name': 'France', 'population': 64, 'currency': 'Euro (€)',
     'tld': 'fr', 'cc': 33},
    {'name': 'Netherlands', 'capital': 'Amsterdam', 'cc': '31'},
    {'name': 'Austria', 'cc': 43, 'currency': 'Euro (€)',
     'population': 8}
]


@templates.test
def as_html():
    table = CountryTable(MEMORY_DATA)
    root = parse(table.as_html())
    assert len(root.findall('.//thead/tr')) == 1
    assert len(root.findall('.//thead/tr/th')) == 4
    assert len(root.findall('.//tbody/tr')) == 4
    assert len(root.findall('.//tbody/tr/td')) == 16

    # no data with no empty_text
    table = CountryTable([])
    root = parse(table.as_html())
    assert 1 == len(root.findall('.//thead/tr'))
    assert 4 == len(root.findall('.//thead/tr/th'))
    assert 0 == len(root.findall('.//tbody/tr'))

    # no data WITH empty_text
    table = CountryTable([], empty_text='this table is empty')
    root = parse(table.as_html())
    assert 1 == len(root.findall('.//thead/tr'))
    assert 4 == len(root.findall('.//thead/tr/th'))
    assert 1 == len(root.findall('.//tbody/tr'))
    assert 1 == len(root.findall('.//tbody/tr/td'))
    assert int(root.find('.//tbody/tr/td').attrib['colspan']) == len(root.findall('.//thead/tr/th'))
    assert root.find('.//tbody/tr/td').text == 'this table is empty'

    # with custom template
    table = CountryTable([], template="django_tables2/table.html")
    table.as_html()


@templates.test
def custom_rendering():
    """For good measure, render some actual templates."""
    countries = CountryTable(MEMORY_DATA)
    context = Context({'countries': countries})

    # automatic and manual column verbose names
    template = Template('{% for column in countries.columns %}{{ column }}/'
                        '{{ column.name }} {% endfor %}')
    result = ('Name/name Capital/capital Population Size/population '
              'Phone Ext./calling_code ')
    assert result == template.render(context)

    # row values
    template = Template('{% for row in countries.rows %}{% for value in row %}'
                        '{{ value }} {% endfor %}{% endfor %}')
    result = ('Germany Berlin 83 49 France — 64 33 Netherlands Amsterdam '
              '— 31 Austria — 8 43 ')
    assert result == template.render(context)


@templates.test
def render_table_templatetag():
    # ensure it works with a multi-order-by
    request = build_request('/')
    table = CountryTable(MEMORY_DATA, order_by=('name', 'population'))
    RequestConfig(request).configure(table)
    template = Template('{% load django_tables2 %}{% render_table table %}')
    html = template.render(Context({'request': request, 'table': table}))

    root = parse(html)
    assert len(root.findall('.//thead/tr')) == 1
    assert len(root.findall('.//thead/tr/th')) == 4
    assert len(root.findall('.//tbody/tr')) == 4
    assert len(root.findall('.//tbody/tr/td')) == 16
    assert root.find('ul[@class="pagination"]/li[@class="cardinality"]').text == '4 items'

    # no data with no empty_text
    table = CountryTable([])
    template = Template('{% load django_tables2 %}{% render_table table %}')
    html = template.render(Context({'request': build_request('/'), 'table': table}))
    root = parse(html)
    assert len(root.findall('.//thead/tr')) == 1
    assert len(root.findall('.//thead/tr/th')) == 4
    assert len(root.findall('.//tbody/tr')) == 0

    # no data WITH empty_text
    request = build_request('/')
    table = CountryTable([], empty_text='this table is empty')
    RequestConfig(request).configure(table)
    template = Template('{% load django_tables2 %}{% render_table table %}')
    html = template.render(Context({'request': request, 'table': table}))
    root = parse(html)
    assert len(root.findall('.//thead/tr')) == 1
    assert len(root.findall('.//thead/tr/th')) == 4
    assert len(root.findall('.//tbody/tr')) == 1
    assert len(root.findall('.//tbody/tr/td')) == 1
    assert int(root.find('.//tbody/tr/td').attrib['colspan']) == len(root.findall('.//thead/tr/th'))
    assert root.find('.//tbody/tr/td').text == 'this table is empty'

    # variable that doesn't exist (issue #8)
    template = Template('{% load django_tables2 %}'
                        '{% render_table this_doesnt_exist %}')
    with raises(ValueError):
        with settings(DEBUG=True):
            template.render(Context())

    # Should still be noisy with debug off
    with raises(ValueError):
        with settings(DEBUG=False):
            template.render(Context())


@templates.test
def render_table_should_support_template_argument():
    table = CountryTable(MEMORY_DATA, order_by=('name', 'population'))
    template = Template('{% load django_tables2 %}'
                        '{% render_table table "dummy.html" %}')
    request = build_request('/')
    context = RequestContext(request, {'table': table})
    assert template.render(context) == 'dummy template contents\n'


@templates.test
def render_table_supports_queryset():
    with database():
        for name in ("Mackay", "Brisbane", "Maryborough"):
            Region.objects.create(name=name)
        template = Template('{% load django_tables2 %}{% render_table qs %}')
        html = template.render(Context({'qs': Region.objects.all(),
                                        'request': build_request('/')}))

        root = parse(html)
        assert [e.text for e in root.findall('.//thead/tr/th/a')] == ["ID", "Name", "Mayor"]
        td = [[td.text for td in tr.findall('td')] for tr in root.findall('.//tbody/tr')]
        db = []
        for region in Region.objects.all():
            db.append([six.text_type(region.id), region.name, "—"])
        assert td == db


@templates.test
def querystring_templatetag():
    template = Template('{% load django_tables2 %}'
                        '<b>{% querystring "name"="Brad" foo.bar=value %}</b>')

    # Should be something like: <root>?name=Brad&amp;a=b&amp;c=5&amp;age=21</root>
    xml = template.render(Context({
        "request": build_request('/?a=b&name=dog&c=5'),
        "foo": {"bar": "age"},
        "value": 21,
    }))

    # Ensure it's valid XML, retrieve the URL
    url = parse(xml).text

    qs = parse_qs(url[1:])  # everything after the ? pylint: disable=C0103
    assert qs["name"] == ["Brad"]
    assert qs["age"] == ["21"]
    assert qs["a"] == ["b"]
    assert qs["c"] == ["5"]


@templates.test
def querystring_templatetag_requires_request():
    with raises(ImproperlyConfigured):
        (Template('{% load django_tables2 %}{% querystring "name"="Brad" %}')
         .render(Context()))


@templates.test
def querystring_templatetag_supports_without():
    context = Context({
        "request": build_request('/?a=b&name=dog&c=5'),
        "a_var": "a",
    })

    template = Template('{% load django_tables2 %}'
                        '<b>{% querystring "name"="Brad" without a_var %}</b>')
    url = parse(template.render(context)).text
    qs = parse_qs(url[1:])  # trim the ? pylint: disable=C0103
    assert set(qs.keys()) == set(["name", "c"])

    # Try with only exclusions
    template = Template('{% load django_tables2 %}'
                        '<b>{% querystring without "a" "name" %}</b>')
    url = parse(template.render(context)).text
    qs = parse_qs(url[1:])  # trim the ? pylint: disable=C0103
    assert set(qs.keys()) == set(["c"])


@templates.test
def title_should_only_apply_to_words_without_uppercase_letters():
    expectations = {
        "a brown fox": "A Brown Fox",
        "a brown foX": "A Brown foX",
        "black FBI": "Black FBI",
        "f.b.i": "F.B.I",
        "start 6pm": "Start 6pm",
    }

    for raw, expected in expectations.items():
        template = Template("{% load django_tables2 %}{{ x|title }}")
        assert template.render(Context({"x": raw})) == expected


@templates.test
def nospaceless_works():
    template = Template("{% load django_tables2 %}"
                        "{% spaceless %}<b>a</b> <i>b {% nospaceless %}<b>c</b>  <b>d</b> {% endnospaceless %}lic</i>{% endspaceless %}")
    assert template.render(Context()) == "<b>a</b><i>b <b>c</b>&#32;<b>d</b> lic</i>"


@templates.test
def whitespace_is_preserved():
    class TestTable(tables.Table):
        name = tables.Column(verbose_name=mark_safe("<b>foo</b> <i>bar</i>"))

    html = TestTable([{"name": mark_safe("<b>foo</b> <i>bar</i>")}]).as_html()

    tree = parse(html)

    assert "<b>foo</b> <i>bar</i>" in lxml.etree.tostring(tree.findall('.//thead/tr/th')[0], encoding='unicode')
    assert "<b>foo</b> <i>bar</i>" in lxml.etree.tostring(tree.findall('.//tbody/tr/td')[0], encoding='unicode')


@templates.test
def as_html_db_queries():
    with database():
        class PersonTable(tables.Table):
            class Meta:
                model = Person

        with queries(count=1):
            PersonTable(Person.objects.all()).as_html()


@templates.test
def render_table_db_queries():
    with database():
        Person.objects.create(first_name="brad", last_name="ayers")
        Person.objects.create(first_name="stevie", last_name="armstrong")

        class PersonTable(tables.Table):
            class Meta:
                model = Person
                per_page = 1

        with queries(count=2):
            # one query for pagination: .count()
            # one query for page records: .all()[start:end]
            request = build_request('/')
            table = PersonTable(Person.objects.all())
            RequestConfig(request).configure(table)
            # render
            (Template('{% load django_tables2 %}{% render_table table %}')
             .render(Context({'table': table, 'request': request})))


@templates.test
def as_csv():
    table = CountryTable(MEMORY_DATA)
    import csv

    fp = StringIO()
    table.as_csv(fp, include_header=False)

    fp.seek(0)
    reader = csv.reader(fp)

    csv_record_count = 0
    for row in reader:
        csv_record_count += 1
    assert len(table.rows) == csv_record_count
    fp.close()


@templates.test
def as_csv_header():
    table = CountryTable(MEMORY_DATA)
    import csv

    fp = StringIO()
    table.as_csv(fp)

    fp.seek(0)
    reader = csv.reader(fp)
    csv_header = reader.next()
    table_column_names = map(lambda c: c.header, table.columns)
    for column in csv_header:
        assert column in table_column_names
    fp.close()


@templates.test
def as_csv_header_exclude_columns():
    table = CountryTable(MEMORY_DATA, exclude=['name'])
    import csv

    fp = StringIO()
    table.as_csv(fp)

    fp.seek(0)
    reader = csv.reader(fp)
    csv_header = reader.next()
    assert 'Name' not in csv_header
    row = reader.next()
    assert len(row) == len(csv_header)
    fp.close()


@templates.test
def as_csv_rows():
    table = CountryTable([MEMORY_DATA[0]])
    import csv

    fp = StringIO()
    table.as_csv(fp)

    fp.seek(0)
    reader = csv.DictReader(fp)
    csv_row = reader.next()
    table_row = table.rows[0]

    assert table_row._get_value('name') == csv_row['Name']
    assert table_row._get_value('calling_code') == int(csv_row['Phone Ext.'])
    assert table_row._get_value('population') == int(csv_row['Population Size'])

    with raises(KeyError):
        # non visible column
        csv_row['Currency']
    fp.close()


@templates.test
def as_excel():
    table = CountryTable(MEMORY_DATA)

    fp = StringIO()
    table.as_excel(fp, include_header=False)

    fp.seek(0)
    from openpyxl import load_workbook
    wb = load_workbook(fp)
    sheet = wb.get_sheet_by_name('Sheet 1')

    assert len(table.rows) == len(sheet.rows)
    fp.close()


@templates.test
def as_excel_header():
    table = CountryTable(MEMORY_DATA)

    fp = StringIO()
    table.as_excel(fp, include_header=True)

    fp.seek(0)
    from openpyxl import load_workbook
    wb = load_workbook(fp)
    sheet = wb.get_sheet_by_name('Sheet 1')
    assert len(table.rows) == len(sheet.rows) - 1

    table_column_names = map(lambda c: c.header, table.columns)

    for column in sheet.rows[0]:
        assert column.value in table_column_names
    fp.close()


@templates.test
def as_excel_sheet_name():
    table = CountryTable(MEMORY_DATA)

    fp = StringIO()
    table.as_excel(fp, title_sheet='My Sheet', include_header=True)

    fp.seek(0)
    from openpyxl import load_workbook
    wb = load_workbook(fp)
    sheet = wb.get_sheet_by_name('My Sheet')
    assert len(table.rows) == len(sheet.rows) - 1

    table_column_names = map(lambda c: c.header, table.columns)

    for column in sheet.rows[0]:
        assert column.value in table_column_names
    fp.close()


@templates.test
def as_excel_header_exclude_columns():
    table = CountryTable(MEMORY_DATA, exclude=['name'])

    fp = StringIO()
    table.as_excel(fp, include_header=True)

    fp.seek(0)
    from openpyxl import load_workbook
    wb = load_workbook(fp)
    sheet = wb.get_sheet_by_name('Sheet 1')
    assert 'Name' not in [c.value for c in sheet.rows[0]]
    assert len(sheet.rows[0]) == len(sheet.rows[1])
    fp.close()


@templates.test
def as_excel_rows():
    table = CountryTable([MEMORY_DATA[0]])

    fp = StringIO()
    table.as_excel(fp, include_header=True)

    fp.seek(0)
    from openpyxl import load_workbook
    wb = load_workbook(fp)
    sheet = wb.get_sheet_by_name('Sheet 1')

    def get_value_under_col_header(header_val):
        col_index = None
        for i, cell in enumerate(sheet.rows[0]):
            if cell.value == header_val:
                col_index = i
        if col_index is None:
            raise KeyError
        return sheet.rows[1][col_index].value

    table_row = table.rows[0]

    assert table_row._get_value('name') == get_value_under_col_header('Name')
    assert table_row._get_value('calling_code') == int(get_value_under_col_header('Phone Ext.'))
    assert table_row._get_value('population') == int(get_value_under_col_header('Population Size'))

    with raises(KeyError):
        # non visible column
        get_value_under_col_header('Currency')
    fp.close()


@templates.test_if(django.VERSION >= (1, 3))
def localization_check():
    def get_cond_localized_table(localizeit=None):
        '''
        helper function for defining Table class conditionally
        '''
        class TestTable(tables.Table):
            name = tables.Column(verbose_name="my column", localize=localizeit)
        return TestTable

    simple_test_data = [{'name': 1234.5}]
    expected_reults = {
        None: '1234.5',
        False: '1234.5',
        True:  '1 234,5'  # non-breaking space
    }

    # no localization
    html = get_cond_localized_table(None)(simple_test_data).as_html()
    assert '<td class="name">{0}</td>'.format(expected_reults[None]) in html

    # unlocalize
    html = get_cond_localized_table(False)(simple_test_data).as_html()
    assert '<td class="name">{0}</td>'.format(expected_reults[False]) in html

    with settings(USE_L10N=True, USE_THOUSAND_SEPARATOR=True):
        with translation("pl"):
            # with default polish locales and enabled thousand separator
            # 1234.5 is formatted as "1 234,5" with nbsp
            html = get_cond_localized_table(True)(simple_test_data).as_html()
            assert '<td class="name">{0}</td>'.format(expected_reults[True]) in html

            # with localize = False there should be no formatting
            html = get_cond_localized_table(False)(simple_test_data).as_html()
            assert '<td class="name">{0}</td>'.format(expected_reults[False]) in html

            # with localize = None and USE_L10N = True
            # there should be the same formatting as with localize = True
            html = get_cond_localized_table(None)(simple_test_data).as_html()
            assert '<td class="name">{0}</td>'.format(expected_reults[True]) in html


@templates.test_if(django.VERSION >= (1, 3))
def localization_check_in_meta():
    class TableNoLocalize(tables.Table):
        name = tables.Column(verbose_name="my column")

        class Meta:
            default = "---"

    class TableLocalize(tables.Table):
        name = tables.Column(verbose_name="my column")

        class Meta:
            default = "---"
            localize = ('name',)

    class TableUnlocalize(tables.Table):
        name = tables.Column(verbose_name="my column")

        class Meta:
            default = "---"
            unlocalize = ('name',)

    class TableLocalizePrecedence(tables.Table):
        name = tables.Column(verbose_name="my column")

        class Meta:
            default = "---"
            unlocalize = ('name',)
            localize = ('name',)

    simple_test_data = [{'name': 1234.5}]
    expected_reults = {
        None: '1234.5',
        False: '1234.5',
        True:  '1{0}234,5'.format(' ')  # non-breaking space
    }

    # No localize
    html = TableNoLocalize(simple_test_data).as_html()
    assert '<td class="name">{0}</td>'.format(expected_reults[None]) in html

    with settings(USE_L10N=True, USE_THOUSAND_SEPARATOR=True):
        with translation("pl"):
            # the same as in localization_check.
            # with localization and polish locale we get formatted output
            html = TableNoLocalize(simple_test_data).as_html()
            assert '<td class="name">{0}</td>'.format(expected_reults[True]) in html

            # localize
            html = TableLocalize(simple_test_data).as_html()
            assert '<td class="name">{0}</td>'.format(expected_reults[True]) in html

            # unlocalize
            html = TableUnlocalize(simple_test_data).as_html()
            assert '<td class="name">{0}</td>'.format(expected_reults[False]) in html

            # test unlocalize higher precedence
            html = TableLocalizePrecedence(simple_test_data).as_html()
            assert '<td class="name">{0}</td>'.format(expected_reults[False]) in html
